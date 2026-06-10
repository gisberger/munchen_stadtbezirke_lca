#!/usr/bin/env python3
"""
Build a Bezirk x Wirkungskategorie impact matrix (per Fahrt) for Munich.

Pipeline
--------
1. Read modal-split + drivetrain + OEPNV-submode shares per Bezirk from data.xlsx.
2. Read the per-km LCA result vector (Impacts sheet, "Result" column) from each
   green leaf-mode file.
3. Decompose every Bezirk into 10 leaf modes and weight each leaf share by its
   per-km LCA vector, then x km/Fahrt  ->  impact per average trip.

Everything is written as LIVE Excel formulas (SUMPRODUCT across sheets), so the
workbook recalculates when you change a share or an LCA value.  Drop the
remaining nine green files into LCA_DIR and re-run to fill the matrix.

Usage:  python build_matrix.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------
DATA_FILE = "data.xlsx"
DATA_SHEET = "Nutzungsstatistik"
LCA_DIR    = Path("lca")
OUTPUT     = "district_impact_matrix.xlsx"

# Leaf mode  ->  (display label, expected green LCA filename)
# Edit the filenames to match your exports; resolution also tries a loose match.
MODE_MAP = {
    "fuss":         ("Fuss",          "fuss_green.xlsx"),
    "rad":          ("Rad",           "rad_green.xlsx"),
    "opnv_ubahn":   ("OEPNV U-Bahn",  "ubahn_green.xlsx"),
    "opnv_sbahn":   ("OEPNV S-Bahn",  "sbahn_gruen.xlsx"),
    "opnv_bus":     ("OEPNV Bus",     "bus_green.xlsx"),
    "opnv_tram":    ("OEPNV Tram",    "tram_green.xlsx"),
    "auto_benzin":  ("Auto Benzin",   "auto_benzin_green.xlsx"),
    "auto_diesel":  ("Auto Diesel",   "auto_diesel_green.xlsx"),
    "auto_phev":    ("Auto PHEV",     "auto_benzin_hybrid_green.xlsx"),
    "auto_elektro": ("Auto Elektro",  "auto_elektro_green.xlsx"),
}
LEAVES = list(MODE_MAP.keys())            # canonical leaf order (shares & LCA share it)

# Loose-match tokens used only if the exact filename is missing.
# (any_of -> at least one must appear; none_of -> must NOT appear, to disambiguate)
TOKENS = {
    "fuss":         (["fuss"],              []),
    "rad":          (["rad"],               ["leihrad"]),
    "opnv_ubahn":   (["ubahn", "u-bahn"],   []),
    "opnv_sbahn":   (["sbahn", "s-bahn"],   []),
    "opnv_bus":     (["bus"],               []),
    "opnv_tram":    (["tram"],              []),
    "auto_benzin":  (["benzin"],            ["hybrid", "phev"]),  # pure gasoline only
    "auto_diesel":  (["diesel"],            []),
    "auto_phev":    (["hybrid", "phev"],    []),
    "auto_elektro": (["elektro", "electric"], []),
}

# Raw columns we pull from data.xlsx
SPLIT_COLS = ["fuss_bezirk", "rad_bezirk", "opnv_bezirk", "auto_bezirk"]
OPNV_COLS  = ["anteil_ubahn_gewichtet", "anteil_sbahn_gewichtet",
              "anteil_bus_gewichtet", "anteil_tram_gewichtet"]
DRIVE_COLS = ["antrieb_benzin_pct", "antrieb_diesel_pct",
              "antrieb_phev_pct", "antrieb_elektro_pct"]
KM_COL     = "entfernung_km_pro_weg"

# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------
FONT       = "Arial"
BLUE       = Font(name=FONT, color="0000FF")          # hardcoded inputs
GREEN      = Font(name=FONT, color="008000")          # cross-sheet formula links
BLACK      = Font(name=FONT, color="000000")
HDR        = Font(name=FONT, bold=True, color="FFFFFF")
TITLE      = Font(name=FONT, bold=True, size=12, color="1F4E78")
HDR_FILL   = PatternFill("solid", fgColor="1F4E78")
SUB_FILL   = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL  = PatternFill("solid", fgColor="FFF2CC")
thin       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
SCI        = '0.000E+00;[Red](0.000E+00);"-"'
PCT        = '0.0%'
NUM1       = '0.0'

def style_hdr(c):
    c.font = HDR; c.fill = HDR_FILL; c.alignment = Alignment("center", "center", wrap_text=True); c.border = BORDER

# ----------------------------------------------------------------------------
# Readers
# ----------------------------------------------------------------------------
def read_districts():
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb[DATA_SHEET]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    need = ["bezirk_nr", "bezirk_name", KM_COL] + SPLIT_COLS + OPNV_COLS + DRIVE_COLS
    missing = [n for n in need if n not in idx]
    if missing:
        raise SystemExit(f"data.xlsx is missing columns: {missing}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[idx["bezirk_nr"]] is None:
            continue
        rows.append({n: r[idx[n]] for n in need})
    wb.close()
    return rows

def resolve_file(key):
    exact = LCA_DIR / MODE_MAP[key][1]
    if exact.exists():
        return exact
    any_of, none_of = TOKENS[key]
    for p in sorted(LCA_DIR.glob("*.xlsx")):
        low = p.name.lower()
        if "green" not in low:
            continue
        if any(t in low for t in any_of) and not any(t in low for t in none_of):
            return p
    return None

def read_impacts(path):
    """Return dict uuid -> (name, unit, per_km_result), preserving order."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Impacts"]
    out = {}
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not header_seen:
            if any(c == "Impact category UUID" for c in cells if isinstance(c, str)):
                header_seen = True
            continue
        # data rows: [None, uuid, name, unit, result, ...]
        if len(cells) < 5 or cells[1] is None:
            continue
        uuid, name, unit, result = cells[1], cells[2], cells[3], cells[4]
        if uuid is None or result is None:
            continue
        out[uuid] = (str(name).strip(), str(unit).strip() if unit else "", float(result))
    wb.close()
    return out

# ----------------------------------------------------------------------------
# Assemble
# ----------------------------------------------------------------------------
districts = read_districts()
print(f"Bezirke gelesen: {len(districts)}")

lca = {}      # leaf -> {uuid: (name, unit, per_km)}
found, missing = [], []
for key in LEAVES:
    p = resolve_file(key)
    if p:
        lca[key] = read_impacts(p)
        found.append((key, p.name, len(lca[key])))
    else:
        lca[key] = {}
        missing.append(key)
print("Gefunden:", [f"{k}<-{f}" for k, f, n in found])
print("Fehlend :", missing)

# canonical category order = from the first non-empty leaf
canon = next((lca[k] for k in LEAVES if lca[k]), {})
if not canon:
    raise SystemExit("No green LCA files found in LCA_DIR.")
cat_uuids = list(canon.keys())
cat_name  = {u: canon[u][0] for u in cat_uuids}
cat_unit  = {u: canon[u][1] for u in cat_uuids}
NCAT = len(cat_uuids)
print(f"Wirkungskategorien: {NCAT}")

# ----------------------------------------------------------------------------
# Workbook
# ----------------------------------------------------------------------------
wb = openpyxl.Workbook()

# ---- Sheet: Info -----------------------------------------------------------
info = wb.active; info.title = "Info"
info["A1"] = "Mobilitaets-LCA Muenchen  -  Bezirk x Wirkungskategorie (pro Fahrt)"
info["A1"].font = TITLE
notes = [
    "",
    "Berechnung je Zelle:  Wirkung_pro_Fahrt = ( SUMME ueber Leaf-Modi: Anteil x LCA_pro_km ) x km_pro_Fahrt",
    "Anteile stammen aus data.xlsx (Bezirks-Modal-Split, OEPNV-Gewichtung, Antriebs-Mix).",
    "LCA-Werte = 'Result' (pro 1 km) aus dem Blatt 'Impacts' der gruenen openLCA-Exporte.",
    "Blau = harte Eingaben | Gruen = Querverweise/Formeln. Aenderungen rechnen automatisch nach.",
    "",
    "Leaf-Modi (Reihenfolge in 'Anteile' und 'LCA_pro_km' identisch):",
]
r = 3
for line in notes:
    info[f"A{r}"] = line; info[f"A{r}"].font = Font(name=FONT); r += 1
for key in LEAVES:
    label, fname = MODE_MAP[key]
    status = "OK" if lca[key] else "FEHLT"
    info[f"A{r}"] = f"  - {label}"
    info[f"B{r}"] = fname
    info[f"C{r}"] = status
    info[f"A{r}"].font = Font(name=FONT)
    info[f"B{r}"].font = Font(name=FONT, italic=True, color="808080")
    info[f"C{r}"].font = Font(name=FONT, bold=True,
                              color="008000" if lca[key] else "C00000")
    if not lca[key]:
        for col in "ABC": info[f"{col}{r}"].fill = WARN_FILL
    r += 1
info.column_dimensions["A"].width = 60
info.column_dimensions["B"].width = 34
info.column_dimensions["C"].width = 8

# ---- Sheet: Rohdaten (blue inputs) -----------------------------------------
roh = wb.create_sheet("Rohdaten")
roh_hdr = (["Nr", "Bezirk"] + SPLIT_COLS + OPNV_COLS + DRIVE_COLS + [KM_COL])
roh.append(roh_hdr)
for c in roh[1]:
    style_hdr(c)
for d in districts:
    roh.append(
        [d["bezirk_nr"], d["bezirk_name"]]
        + [d[c] for c in SPLIT_COLS] + [d[c] for c in OPNV_COLS]
        + [d[c] for c in DRIVE_COLS] + [d[KM_COL]]
    )
for row in roh.iter_rows(min_row=2):
    for c in row:
        c.border = BORDER
        if c.column >= 3:                      # numeric inputs
            c.font = BLUE
        else:
            c.font = BLACK
roh.column_dimensions["A"].width = 5
roh.column_dimensions["B"].width = 30
for i in range(3, len(roh_hdr) + 1):
    roh.column_dimensions[get_column_letter(i)].width = 13
roh.freeze_panes = "C2"
# column letters in Rohdaten
R = {name: get_column_letter(roh_hdr.index(name) + 1) for name in roh_hdr}

# ---- Sheet: LCA_pro_km (blue inputs, categories in ROWS, modes in COLS) -----
lcas = wb.create_sheet("LCA_pro_km")
lcas["A1"] = "LCA pro km (gruen)"; lcas["A1"].font = TITLE
# header row 2
lcas.cell(2, 1, "Wirkungskategorie"); style_hdr(lcas.cell(2, 1))
for j, key in enumerate(LEAVES):
    c = lcas.cell(2, 2 + j, MODE_MAP[key][0]); style_hdr(c)
unit_col = 2 + len(LEAVES)
lcas.cell(2, unit_col, "Einheit"); style_hdr(lcas.cell(2, unit_col))
# data rows 3..
for i, u in enumerate(cat_uuids):
    rr = 3 + i
    nc = lcas.cell(rr, 1, cat_name[u]); nc.font = BLACK; nc.border = BORDER
    for j, key in enumerate(LEAVES):
        val = lca[key].get(u)
        cell = lcas.cell(rr, 2 + j)
        if val is not None:
            cell.value = val[2]; cell.number_format = SCI
        cell.font = BLUE; cell.border = BORDER
    uc = lcas.cell(rr, unit_col, cat_unit[u]); uc.font = Font(name=FONT, italic=True, color="808080"); uc.border = BORDER
lcas.column_dimensions["A"].width = 42
for j in range(len(LEAVES)):
    lcas.column_dimensions[get_column_letter(2 + j)].width = 13
lcas.column_dimensions[get_column_letter(unit_col)].width = 18
lcas.freeze_panes = "B3"
LCA_FIRST = get_column_letter(2)                       # B
LCA_LAST  = get_column_letter(1 + len(LEAVES))         # K (10 modes)

# ---- Sheet: Anteile (green formulas from Rohdaten) --------------------------
sh = wb.create_sheet("Anteile")
sh_hdr = ["Nr", "Bezirk"] + [MODE_MAP[k][0] for k in LEAVES] + ["Summe", "km/Fahrt"]
sh.append(sh_hdr)
for c in sh[1]:
    style_hdr(c)
SH_FIRST = get_column_letter(3)                         # C  (first leaf share)
SH_LAST  = get_column_letter(2 + len(LEAVES))           # L  (last leaf share)
for i, d in enumerate(districts):
    r = 2 + i           # Anteile data row
    rr = 2 + i          # Rohdaten data row (aligned)
    sh.cell(r, 1, d["bezirk_nr"]).font = BLACK
    sh.cell(r, 2, d["bezirk_name"]).font = BLACK
    # leaf share formulas (fractions 0..1)
    f = {
        "fuss":         f"=Rohdaten!{R['fuss_bezirk']}{rr}/100",
        "rad":          f"=Rohdaten!{R['rad_bezirk']}{rr}/100",
        "opnv_ubahn":   f"=Rohdaten!{R['opnv_bezirk']}{rr}/100*Rohdaten!{R['anteil_ubahn_gewichtet']}{rr}/100",
        "opnv_sbahn":   f"=Rohdaten!{R['opnv_bezirk']}{rr}/100*Rohdaten!{R['anteil_sbahn_gewichtet']}{rr}/100",
        "opnv_bus":     f"=Rohdaten!{R['opnv_bezirk']}{rr}/100*Rohdaten!{R['anteil_bus_gewichtet']}{rr}/100",
        "opnv_tram":    f"=Rohdaten!{R['opnv_bezirk']}{rr}/100*Rohdaten!{R['anteil_tram_gewichtet']}{rr}/100",
        "auto_benzin":  f"=Rohdaten!{R['auto_bezirk']}{rr}/100*Rohdaten!{R['antrieb_benzin_pct']}{rr}/100",
        "auto_diesel":  f"=Rohdaten!{R['auto_bezirk']}{rr}/100*Rohdaten!{R['antrieb_diesel_pct']}{rr}/100",
        "auto_phev":    f"=Rohdaten!{R['auto_bezirk']}{rr}/100*Rohdaten!{R['antrieb_phev_pct']}{rr}/100",
        "auto_elektro": f"=Rohdaten!{R['auto_bezirk']}{rr}/100*Rohdaten!{R['antrieb_elektro_pct']}{rr}/100",
    }
    for j, key in enumerate(LEAVES):
        c = sh.cell(r, 3 + j, f[key]); c.font = GREEN; c.number_format = PCT; c.border = BORDER
    summ = sh.cell(r, 3 + len(LEAVES), f"=SUM({SH_FIRST}{r}:{SH_LAST}{r})")
    summ.font = BLACK; summ.number_format = PCT; summ.border = BORDER
    km = sh.cell(r, 4 + len(LEAVES), f"=Rohdaten!{R[KM_COL]}{rr}")
    km.font = GREEN; km.number_format = NUM1; km.border = BORDER
SH_SUM_COL = get_column_letter(3 + len(LEAVES))         # M
SH_KM_COL  = get_column_letter(4 + len(LEAVES))         # N
sh.column_dimensions["A"].width = 5
sh.column_dimensions["B"].width = 30
for i in range(3, 4 + len(LEAVES) + 1):
    sh.column_dimensions[get_column_letter(i)].width = 11
sh.freeze_panes = "C2"

# ---- Sheet: Matrix_pro_Fahrt (green SUMPRODUCT formulas) --------------------
mx = wb.create_sheet("Matrix_pro_Fahrt")
mx["A1"] = "Wirkung pro Fahrt  -  Bezirk x Wirkungskategorie (gruener Strom)"
mx["A1"].font = TITLE
mx.cell(2, 1, "Nr");     style_hdr(mx.cell(2, 1))
mx.cell(2, 2, "Bezirk"); style_hdr(mx.cell(2, 2))
mx.cell(3, 1, "");       mx.cell(3, 2, "Einheit").font = Font(name=FONT, bold=True, italic=True)
for k, u in enumerate(cat_uuids):
    col = 3 + k
    hc = mx.cell(2, col, cat_name[u]); style_hdr(hc)
    uc = mx.cell(3, col, cat_unit[u]); uc.font = Font(name=FONT, italic=True, color="808080")
    uc.fill = SUB_FILL; uc.alignment = Alignment("center"); uc.border = BORDER
for i, d in enumerate(districts):
    r = 4 + i                      # matrix data row
    sr = 2 + i                     # matching Anteile row
    mx.cell(r, 1, d["bezirk_nr"]).font = BLACK
    mx.cell(r, 2, d["bezirk_name"]).font = BLACK
    mx.cell(r, 1).border = BORDER; mx.cell(r, 2).border = BORDER
    for k in range(NCAT):
        col = 3 + k
        lca_row = 3 + k            # category row in LCA_pro_km
        formula = (f"=SUMPRODUCT(Anteile!${SH_FIRST}{sr}:${SH_LAST}{sr},"
                   f"LCA_pro_km!${LCA_FIRST}{lca_row}:${LCA_LAST}{lca_row})"
                   f"*Anteile!${SH_KM_COL}{sr}")
        c = mx.cell(r, col, formula)
        c.font = GREEN; c.number_format = SCI; c.border = BORDER
mx.column_dimensions["A"].width = 5
mx.column_dimensions["B"].width = 30
for k in range(NCAT):
    mx.column_dimensions[get_column_letter(3 + k)].width = 13
mx.freeze_panes = "C4"

wb.save(OUTPUT)
print(f"\nGespeichert: {OUTPUT}")
print(f"  Matrix: {len(districts)} Bezirke x {NCAT} Kategorien (pro Fahrt)")
